import pandas as pd
from django.db.models import Case, When, CharField, Value

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO


from django.http import HttpResponse

from abc import ABC
from django.db.models import Q
from django.db import transaction
from commons.middlewares.exception import BadRequestException

from staffing_status.models import StaffingStatus
from personnel_database.models.subsatker import SubSatKer

CUSTOM_ORDER = [
    "PIMPINAN",
    "DIT KAMSEL",
    "DIT GAKKUM",
    "DIT REGIDENT",
    "BAG OPS",
    "BAG RENMIN",
    "BAG TIK",
    "SIKEU",
    "TAUD"
]
ORDER_DICT = {name: i for i, name in enumerate(CUSTOM_ORDER)}

class StaffingService(ABC):
    
    @classmethod
    @transaction.atomic
    def update_staffing_status(cls, **kwargs) :
        subsatker = kwargs.pop("satker")
        data = kwargs.pop("data")
        staffing_status_list = StaffingStatus.objects.filter(subsatker__nama = subsatker)

        if(len(staffing_status_list) == 0) :
            raise BadRequestException(f"Satker {subsatker} not exists in staffing status.")
        
        for i in data :
            i = dict(i)
            staffing_status = staffing_status_list.filter(Q(nama=i['pangkat']) | Q(pangkat__nama=i['pangkat'])).first()
            if(not staffing_status) :
                raise BadRequestException(f"Pangkat {i['pangkat']} not exists in staffing status.")            
            cls.update_data(staffing_status, i['dsp'])

        return cls.get_staffing_status()

    @classmethod
    def update_data(cls, data: StaffingStatus, dsp: int) :
        data.dsp = dsp
        data.save()

    @classmethod
    def get_staffing_status(cls) :
        data = []

        satker_list = sorted(SubSatKer.objects.all(), key=lambda x: ORDER_DICT.get(x.nama, len(CUSTOM_ORDER)))
        for i in satker_list :
            temp_polri = cls.get_staffing_data("POLRI", i)
            temp_pns_polri = cls.get_staffing_data("PNS POLRI", i)
            temp = cls.data_wrapper(i.nama, temp_polri, temp_pns_polri)
            data.append(temp)

        return data

    @classmethod
    def get_staffing_data(cls, tipe: str, subsatker:SubSatKer) :
        data = {
                tipe: {
                    "jumlah" : {
                        "dsp" : 0,
                        "rill" : 0
                    }
                }}
        
        staffing_status_list = StaffingStatus.objects.filter(Q(pangkat__tipe = tipe) & Q(subsatker=subsatker)).distinct()
        for i in staffing_status_list :
            message = ""
            if(i.dsp > i.rill) :
                message = f"Subsatker {i.subsatker} dengan Pangkat {i.nama} kekurangan {i.dsp - i.rill} personil"
            elif(i.dsp < i.rill) :
                message = f"Subsatker {i.subsatker} dengan Pangkat {i.nama} kelebihan {i.rill - i.dsp} personil"
  
            temp = {
                "dsp" : i.dsp,
                "rill" : i.rill,
                "message" : message
            }
            data[tipe][i.nama] = temp
            data[tipe]['jumlah']['dsp'] = data[tipe]['jumlah']['dsp'] + i.dsp
            data[tipe]['jumlah']['rill'] = data[tipe]['jumlah']['rill'] + i.rill

        return data
    
    @classmethod
    def data_wrapper(cls, satker, polri: dict, pns_polri: dict) :
        data = {"satker" : satker,
                "keterangan" : {
                    "dsp" : polri["POLRI"]["jumlah"]["dsp"] + pns_polri["PNS POLRI"]["jumlah"]["dsp"],
                    "rill" : polri["POLRI"]["jumlah"]["rill"] + pns_polri["PNS POLRI"]["jumlah"]["rill"]
                },
                **polri, **pns_polri
        }
        return data
    
    @classmethod
    def get_total_by_pangkat(cls) :
        data = {
            "POLRI" : {
                "dsp": 0,
                "rill": 0
            },
            "PNS POLRI" : {
                "dsp" : 0,
                "rill" : 0
            },
            "Keterangan" : {
                "dsp" : 0,
                "rill" : 0,
            }
        }

        staffing_status_list = StaffingStatus.objects.all().distinct().annotate(
                tipe=Case(
                When(pangkat__tipe='PNS POLRI', then=Value("PNS POLRI")),
                default=Value("POLRI"),
                output_field=CharField(),
                )
        )

        for i in staffing_status_list :
            if(not data.get(i.nama, None)) :
                data[i.nama] = {
                    'dsp' : 0,
                    'rill' : 0,
                }

            data[i.nama]['dsp'] = data[i.nama]['dsp'] +  i.dsp
            data[i.nama]['rill'] = data[i.nama]['rill'] + i.rill

            data[i.tipe]['dsp'] = data[i.tipe]['dsp'] + i.dsp
            data[i.tipe]['rill'] = data[i.tipe]['rill'] + i.rill

            data['Keterangan']['dsp'] = data['Keterangan']['dsp'] + i.dsp
            data['Keterangan']['rill'] = data['Keterangan']['rill'] + i.rill

        return data
    
    @classmethod
    def export_csv_file(cls):
    # Define the MultiIndex for columns
        # Define the MultiIndex for columns
        columns = pd.MultiIndex.from_tuples([
            ('', 'No'), ('', 'SatKer'),
            ('IRJEN', 'DSP'), ('IRJEN', 'RIIL'),
            ('BRIGJEN', 'DSP'), ('BRIGJEN', 'RIIL'),
            ('KOMBES', 'DSP'), ('KOMBES', 'RIIL'),
            ('AKBP', 'DSP'), ('AKBP', 'RIIL'),
            ('KOMPOL', 'DSP'), ('KOMPOL', 'RIIL'),
            ('AKP', 'DSP'), ('AKP', 'RIIL'),
            ('IP', 'DSP'), ('IP', 'RIIL'),
            ('BRIG/TA', 'DSP'), ('BRIG/TA', 'RIIL'),
            ('Jumlah', 'DSP'), ('Jumlah', 'RIIL'),
            ('IV', 'DSP'), ('IV', 'RIIL'),
            ('III', 'DSP'), ('III', 'RIIL'),
            ('II/I', 'DSP'), ('II/I', 'RIIL'),
            ('Jumlah PNS', 'DSP'), ('Jumlah PNS', 'RIIL'), 
            ('Ket', 'DSP'), ('Ket', 'RIIL')
        ], names=['', ''])
         
        res = dict()

        for staffing_status_obj in StaffingStatus.objects.filter():
            if staffing_status_obj.subsatker.nama not in res:
                res[staffing_status_obj.subsatker.nama] = {}
            
            pangkat = staffing_status_obj.pangkat.all()
            for pangkat_obj in pangkat:
                tipe = pangkat_obj.tipe
                if tipe not in res[staffing_status_obj.subsatker.nama]:
                    res[staffing_status_obj.subsatker.nama][tipe] = {}
                res[staffing_status_obj.subsatker.nama][tipe][staffing_status_obj.nama] = {'dsp':staffing_status_obj.dsp, 'riil':staffing_status_obj.rill}
    

        # Create the DataFrame
        # df = pd.DataFrame(data, columns=columns)
        df = pd.DataFrame(columns=columns)
        sorted_data = sorted(res.items(), key=lambda item: ORDER_DICT.get(item[0].upper(), len(CUSTOM_ORDER)))
        sorted_data_dict = dict(sorted_data)
        res = sorted_data_dict

        for idx, (satker, levels) in enumerate(res.items(), start=1):
            row = [idx, satker]
            polri_data = []
            pns_polri_data = []
            for level in ['IRJEN', 'BRIGJEN', 'KOMBES', 'AKBP', 'KOMPOL', 'AKP', 'IP', 'BRIG/TA']:
                if 'POLRI' in levels and level in levels['POLRI']:
                    polri_data.extend([levels['POLRI'][level]['dsp'], levels['POLRI'][level]['riil']])
                else:
                    polri_data.extend([0, 0])
            polri_data.extend([sum(polri_data[::2]), sum(polri_data[1::2])])
            for level in ['IV', 'III', 'II/I']:
                if 'PNS POLRI' in levels and level in levels['PNS POLRI']:
                    pns_polri_data.extend([levels['PNS POLRI'][level]['dsp'], levels['PNS POLRI'][level]['riil']])
                else:
                    pns_polri_data.extend([0, 0])
            pns_polri_data.extend([sum(pns_polri_data[::2]), sum(pns_polri_data[1::2])])
            row.extend(polri_data)
            row.extend(pns_polri_data)
            row.extend([sum(polri_data[:-2:2]) + sum(pns_polri_data[:-2:2]), sum(polri_data[1:-2:2]) + sum(pns_polri_data[1:-2:2])])
            df.loc[len(df)] = row

        # Calculate totals
        totals = [
            ['Jumlah', ''] + [df.iloc[:, i].sum() for i in range(2, len(df.columns))]
        ]
        # Convert totals to DataFrame
        totals_df = pd.DataFrame(totals, columns=columns)
        # Concatenate the totals row to the original DataFrame
        df = pd.concat([df, totals_df], ignore_index=True)

        # Create an Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active
        # Append the dataframe to the worksheet starting from row 2
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Merge cells for the header
        ws.merge_cells('C2:D2')
        ws.merge_cells('E2:F2')
        ws.merge_cells('G2:H2')
        ws.merge_cells('I2:J2')
        ws.merge_cells('K2:L2')
        ws.merge_cells('M2:N2')
        ws.merge_cells('O2:P2')
        ws.merge_cells('Q2:R2')
        ws.merge_cells('S2:T2')
        ws.merge_cells('U2:V2')
        ws.merge_cells('W2:X2')
        ws.merge_cells('Y2:Z2')
        ws.merge_cells('AA2:AB2')
        ws.merge_cells('AC2:AD2')
        
        ws.merge_cells('C1:T1')
        ws.merge_cells('U1:AB1')

        ws['C1'].value = 'POLRI'
        ws['U1'].value = 'PNS POLRI'

        ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['U1'].alignment = Alignment(horizontal='center', vertical='center')

        headers = ['IRJEN', 'BRIGJEN', 'KOMBES', 'AKBP', 'KOMPOL', 'AKP', 'IP', 'BRIG/TA', 'Jumlah', 'IV', 'III', 'II/I', 'Jumlah', 'Ket']
        for i, header in enumerate(headers, start=3):
            ws.cell(row=2, column=i*2-1-2).value = header
            ws.cell(row=2, column=i*2-1-2).alignment = Alignment(horizontal='center', vertical='center')

        # Align the subheaders
        for col_num, (header, subheader) in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook to a virtual file
        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)

        # Create HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="staffing-status.xlsx"'
        response.write(virtual_workbook.getvalue())

        return response